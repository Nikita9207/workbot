#!/usr/bin/env python3
"""
Конвертер Excel файлов пауэрлифтинг программ в JSON формат
"""
import json
import sys
import os
import re


def parse_sets_reps(value):
    """Парсит формат типа '3*4' (3 подхода по 4 повтора) или просто '5' (1 подход на 5)"""
    if value is None or value == '':
        return None, None

    val_str = str(value).strip()

    # Формат "3*4" или "3x4"
    match = re.match(r'(\d+)\s*[*xх]\s*(\d+)', val_str)
    if match:
        return int(match.group(1)), int(match.group(2))

    # Просто число - 1 подход на N повторов
    if val_str.replace('.0', '').isdigit():
        return 1, int(float(val_str))

    return None, None


def convert_muravyev_xls(filepath, output_path):
    """Конвертирует Tsikl-Muravyeva.xls в JSON"""
    import xlrd

    wb = xlrd.open_workbook(filepath)

    # Проценты из заголовка
    percents = [50, 60, 65, 70, 75, 80, 85]

    program = {
        "name": "Цикл Муравьёва",
        "author": "Муравьёв",
        "level": ["2_разряд", "1_разряд", "КМС"],
        "weeks": 17,
        "days_per_week": 2,
        "exercises": {}
    }

    # Парсим каждый лист (Жим, Присед, Тяга)
    exercise_map = {
        'Жим': 'bench',
        'Присед': 'squat',
        'Тяга': 'deadlift'
    }

    for sheet_name in wb.sheet_names():
        if sheet_name not in exercise_map:
            continue

        ws = wb.sheet_by_name(sheet_name)
        ex_key = exercise_map[sheet_name]

        exercise_data = {
            "name": f"{sheet_name} на 17",
            "weeks": []
        }

        # Читаем данные начиная с row 3 (после заголовков)
        current_week = None

        for row_idx in range(3, ws.nrows):
            row = [ws.cell_value(row_idx, col) for col in range(10)]

            week_cell = row[0]
            training_num = row[1]  # I или II
            day = row[2]  # Пн или Пт

            # Новая неделя
            if week_cell and str(week_cell).strip():
                try:
                    week_num = int(float(week_cell))
                    current_week = week_num
                except:
                    continue

            if current_week is None:
                continue

            # Собираем сеты для этой тренировки
            sets = []
            for col_idx, percent in enumerate(percents):
                cell_value = row[3 + col_idx] if 3 + col_idx < len(row) else None
                if cell_value is None or cell_value == '':
                    continue

                num_sets, reps = parse_sets_reps(cell_value)
                if num_sets and reps:
                    sets.append({
                        "percent": percent,
                        "reps": reps,
                        "sets": num_sets
                    })

            if sets:
                exercise_data["weeks"].append({
                    "week_num": current_week,
                    "sets": sets
                })

        program["exercises"][ex_key] = exercise_data

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(program, f, ensure_ascii=False, indent=2)

    print(f"  Сохранено: {output_path}")
    print(f"  Упражнений: {len(program['exercises'])}")
    return program


def convert_golovinsky_xlsm(filepath, output_path, cycle_num):
    """Конвертирует cycle*.xlsm (Головинский) в JSON"""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)

    program = {
        "name": f"Головинский Цикл {cycle_num}",
        "author": "Головинский",
        "level": ["2_разряд", "1_разряд", "КМС"],
        "weeks": 0,
        "days_per_week": 4,
        "microcycles": []
    }

    # Данные на листе "Цикл"
    ws = wb['Цикл']

    current_micro = None
    current_day = None
    day_mapping = {'пн': 1, 'вт': 2, 'ср': 3, 'чт': 4, 'пт': 5, 'сб': 6, 'вс': 7}

    for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
        if all(cell is None for cell in row):
            continue

        first_cell = str(row[0]).strip().lower() if row[0] else ''

        # Новый микроцикл
        if 'микроцикл' in first_cell:
            match = re.search(r'(\d+)', first_cell)
            if match:
                if current_micro and current_micro["workouts"]:
                    program["microcycles"].append(current_micro)

                micro_num = int(match.group(1))
                current_micro = {
                    "micro_num": micro_num,
                    "workouts": []
                }
                current_day = None
            continue

        # Пропускаем строку заголовка
        if 'дата' in first_cell or 'нагрузка' in first_cell:
            continue

        # День недели
        if first_cell in day_mapping and current_micro is not None:
            current_day = {
                "day": day_mapping[first_cell],
                "exercises": []
            }
            current_micro["workouts"].append(current_day)

        # Упражнение
        if current_day is not None:
            # Колонка 2 - название упражнения
            exercise_name = str(row[2]).strip() if row[2] else ''

            if exercise_name and exercise_name.lower() not in ['none', '']:
                # Колонка 5 - вес, 6 - повторы, 7 - подходы, 8 - процент
                weight = row[5] if len(row) > 5 and row[5] else None
                reps = row[6] if len(row) > 6 and row[6] else None
                sets_count = row[7] if len(row) > 7 and row[7] else 1
                percent = row[8] if len(row) > 8 and row[8] else None

                if weight or reps:
                    exercise = {
                        "name": exercise_name,
                        "sets": []
                    }

                    try:
                        sets_int = int(float(sets_count)) if sets_count else 1
                        reps_int = int(float(reps)) if reps else 0
                        percent_float = float(percent) * 100 if percent and float(percent) < 1 else (float(percent) if percent else 0)

                        if reps_int > 0:
                            exercise["sets"].append({
                                "percent": round(percent_float, 1) if percent_float else None,
                                "reps": reps_int,
                                "sets": sets_int
                            })
                            current_day["exercises"].append(exercise)
                    except (ValueError, TypeError):
                        pass

    if current_micro and current_micro["workouts"]:
        program["microcycles"].append(current_micro)

    program["weeks"] = len(program["microcycles"])

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(program, f, ensure_ascii=False, indent=2)

    print(f"  Сохранено: {output_path}")
    print(f"  Микроциклов: {len(program['microcycles'])}")
    return program


def main():
    books_dir = "/Users/nikitakrasilnikov/Desktop/Книги"
    output_dir = "/Users/nikitakrasilnikov/GolandProjects/workbot/clients/ai/templates"

    print("=== Конвертация Excel в JSON ===\n")

    # 1. Муравьёв
    print("1. Конвертирую Цикл Муравьёва...")
    muravyev_path = os.path.join(books_dir, "Tsikl-Muravyeva.xls")
    if os.path.exists(muravyev_path):
        convert_muravyev_xls(muravyev_path, os.path.join(output_dir, "muravyev_cycle.json"))

    # 2. Головинский Цикл 2
    print("\n2. Конвертирую Головинский Цикл 2...")
    cycle2_path = os.path.join(books_dir, "cycle2.xlsm")
    if os.path.exists(cycle2_path):
        convert_golovinsky_xlsm(cycle2_path, os.path.join(output_dir, "golovinsky_cycle_2.json"), 2)

    # 3. Головинский Цикл 11
    print("\n3. Конвертирую Головинский Цикл 11...")
    cycle11_path = os.path.join(books_dir, "cycle11.xlsm")
    if os.path.exists(cycle11_path):
        convert_golovinsky_xlsm(cycle11_path, os.path.join(output_dir, "golovinsky_cycle_11.json"), 11)

    print("\n=== Готово! ===")


if __name__ == "__main__":
    main()
